"""
Odczyt modelu (np. redistribution / reaccumulation) z PNG TradingView + dopisanie wiersza do Excela.

Wywoływane po każdym udanym zrzucie z harmonogramu (OCR + openpyxl).
Wymaga zainstalowanego Tesseract OCR (Windows: https://github.com/UB-Mannheim/tesseract/wiki).

Kolumny zgodne ze schema_models.xlsx: timestamp, current_model, previous_model,
m5_context_model, h4_context, HH, LH, LL, HL — ostatnie cztery = przyrost nowych etykiet względem poprzedniego zrzutu (puste gdy brak zmiany).
"""
from __future__ import annotations

import json
import logging
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

_LOG = logging.getLogger("tv_models_excel")

_TS_PL = ZoneInfo("Europe/Warsaw")


def _excel_timestamp_cell(timestamp_pl_iso: str) -> datetime:
    """ISO z harmonogramu → naiwny czas ścienny PL (Excel wyświetli jako data/godzina)."""
    d = datetime.fromisoformat(timestamp_pl_iso.replace("Z", "+00:00"))
    if d.tzinfo is not None:
        d = d.astimezone(_TS_PL)
    return d.replace(tzinfo=None, microsecond=0)

# Jak w Twoim schema_models.xlsx
HEADERS: tuple[str, ...] = (
    "timestamp",
    "current_model",
    "previous_model",
    "m5_context_model",
    "h4_context",
    "HH",
    "LH",
    "LL",
    "HL",
)

STRUCT_KEYS = ("HH", "LH", "LL", "HL")

# Modele typu redistribution-2, reaccumulation-1 (OCR bywa z spacjami)
_MODEL_PATTERNS = [
    re.compile(
        r"(?i)\b(re[-\s]*accumulation|re[-\s]*distribution|redistribution|reaccumulation)\b\s*[-]?\s*(\d+)"
    ),
    re.compile(r"(?i)\b(accumulation|distribution)\b\s*[-]?\s*(\d+)"),
]

# Etykiety struktury na wykresie
_LABEL_RE = re.compile(r"\b(HH|LH|LL|HL)\b", re.I)


def _resolve_tesseract_exe(explicit: str | None) -> str | None:
    """PATH → jawna ścieżka z JSON → domyślny Program Files (Windows)."""
    s = (explicit or "").strip()
    if s:
        return s
    w = shutil.which("tesseract")
    if w:
        return w
    if sys.platform == "win32":
        tw = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
        if tw.is_file():
            return str(tw)
    return None


def _fraction_crop_box(
    w: int, h: int, box: list[float] | tuple[float, float, float, float]
) -> tuple[int, int, int, int]:
    x0, y0, x1, y1 = (float(t) for t in box)
    left = max(0, int(w * x0))
    top = max(0, int(h * y0))
    right = min(w, int(w * x1))
    bottom = min(h, int(h * y1))
    if right <= left or bottom <= top:
        return 0, 0, w, h
    return left, top, right, bottom


def _ocr_image(
    path: Path,
    tesseract_cmd: str | None,
    crop_fractions: list[tuple[float, float, float, float]],
    ocr_lang: str = "eng",
) -> str:
    import pytesseract
    from PIL import Image, ImageEnhance

    exe = _resolve_tesseract_exe(tesseract_cmd)
    if not exe:
        raise RuntimeError(
            "Brak Tesseract OCR — zainstaluj (Windows: UB-Mannheim) i dodaj do PATH "
            "lub ustaw models_log.tesseract_cmd w capture_schedule.json"
        )
    pytesseract.pytesseract.tesseract_cmd = exe

    img = Image.open(path).convert("RGB")
    parts: list[str] = []
    w, h = img.size
    for frac in crop_fractions:
        l, t, r, b = _fraction_crop_box(w, h, frac)
        crop_img = img.crop((l, t, r, b))
        g = crop_img.convert("L")
        g = ImageEnhance.Contrast(g).enhance(2.0)
        g = g.resize((max(1, g.width * 2), max(1, g.height * 2)))
        txt = pytesseract.image_to_string(g, lang=ocr_lang, config="--psm 6")
        parts.append(txt)
    return "\n".join(parts)


def _default_crops(cfg: dict[str, Any]) -> list[tuple[float, float, float, float]]:
    leg = cfg.get("ocr_legend_crop", [0.0, 0.06, 0.38, 0.52])
    ch = cfg.get("ocr_chart_crop", [0.0, 0.06, 1.0, 0.94])
    return [tuple(leg), tuple(ch)]  # type: ignore[return-value]


def _parse_model(text: str) -> str:
    best = ""
    for pat in _MODEL_PATTERNS:
        for m in pat.finditer(text):
            g1 = m.group(1).lower().replace(" ", "").replace("-", "")
            g2 = m.group(2) if m.lastindex and m.lastindex >= 2 else ""
            cand = f"{g1}-{g2}" if g2 else g1
            if len(cand) > len(best):
                best = cand
    return best.strip()


def _count_structure_labels(text: str) -> dict[str, int]:
    c = {k: 0 for k in STRUCT_KEYS}
    for m in _LABEL_RE.finditer(text.upper()):
        k = m.group(1).upper()
        if k in c:
            c[k] += 1
    return c


def _load_state(state_path: Path) -> dict[str, Any]:
    if not state_path.is_file():
        return {"version": 1, "by_tf": {}}
    try:
        with state_path.open(encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"version": 1, "by_tf": {}}


def _save_state(state_path: Path, data: dict[str, Any]) -> None:
    state_path.parent.mkdir(parents=True, exist_ok=True)
    with state_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _ensure_sheet(wb: Any, sheet_name: str) -> Any:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(list(HEADERS))
    return ws


def _resolve_paths(repo_root: Path, models_cfg: dict[str, Any]) -> tuple[Path, Path]:
    excel_rel = str(models_cfg.get("excel_path", "models_logs/trading_models.xlsx"))
    p = Path(excel_rel)
    excel_path = p if p.is_absolute() else repo_root / p
    state_name = str(models_cfg.get("state_filename", "ocr_state.json"))
    state_path = excel_path.parent / state_name
    return excel_path, state_path


def log_capture_to_excel(
    *,
    repo_root: Path,
    models_cfg: dict[str, Any],
    tf_label: str,
    png_path: Path,
    timestamp_pl_iso: str,
    parent_log: logging.Logger | None = None,
) -> bool:
    """
    Dopisuje jeden wiersz do arkusza ``{tf_label}_models_log``.
    Zwraca True tylko po udanym zapisie Excel + stanu OCR.
    """
    log = parent_log or _LOG
    if not png_path.is_file():
        log.warning("models_excel: brak PNG %s", png_path)
        return False

    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as e:
        log.warning("models_excel: brak biblioteki %s — pip install openpyxl pytesseract Pillow", e)
        return False

    tesseract_cmd = (models_cfg.get("tesseract_cmd") or "").strip() or None
    crops_cfg = _default_crops(models_cfg)
    ocr_lang = str(models_cfg.get("ocr_lang", "eng") or "eng").strip() or "eng"
    try:
        raw = _ocr_image(png_path, tesseract_cmd, crops_cfg, ocr_lang=ocr_lang)
    except Exception:
        log.exception("models_excel: OCR nieudany tf=%s file=%s", tf_label, png_path)
        return False

    current_model = _parse_model(raw)
    counts = _count_structure_labels(raw)

    excel_path, state_path = _resolve_paths(repo_root, models_cfg)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    sheet_name = f"{tf_label}_models_log"

    if excel_path.is_file():
        wb = load_workbook(excel_path)
    else:
        tpl_raw = models_cfg.get("schema_template")
        tpl_path = None
        if tpl_raw:
            t = Path(str(tpl_raw))
            tpl_path = t if t.is_absolute() else repo_root / t
        if tpl_path is not None and tpl_path.is_file():
            shutil.copy2(tpl_path, excel_path)
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)
            _ensure_sheet(wb, sheet_name)

    ws = _ensure_sheet(wb, sheet_name)

    prev_model = ""
    max_row = ws.max_row
    if max_row >= 2:
        # kolumna B = current_model
        v = ws.cell(row=max_row, column=2).value
        if v is not None:
            prev_model = str(v).strip()

    state = _load_state(state_path)
    by_tf: dict[str, Any] = state.setdefault("by_tf", {})
    st_tf: dict[str, Any] = by_tf.setdefault(
        tf_label, {"structure_counts": {k: 0 for k in STRUCT_KEYS}, "has_baseline": False}
    )
    old_c: dict[str, int] = {k: int(st_tf["structure_counts"].get(k, 0)) for k in STRUCT_KEYS}
    has_base = bool(st_tf.get("has_baseline"))

    hh_v = lh_v = ll_v = hl_v = ""
    if not has_base:
        st_tf["has_baseline"] = True
        for k in STRUCT_KEYS:
            st_tf["structure_counts"][k] = counts[k]
    else:
        for k in STRUCT_KEYS:
            d = counts[k] - old_c.get(k, 0)
            if d > 0:
                if k == "HH":
                    hh_v = f"+{d}"
                elif k == "LH":
                    lh_v = f"+{d}"
                elif k == "LL":
                    ll_v = f"+{d}"
                else:
                    hl_v = f"+{d}"
        for k in STRUCT_KEYS:
            st_tf["structure_counts"][k] = counts[k]

    m5_ctx = str(models_cfg.get("m5_context_default", "") or "")
    h4_ctx = str(models_cfg.get("h4_context_default", "") or "")

    ts_cell = _excel_timestamp_cell(timestamp_pl_iso)
    row = [
        ts_cell,
        current_model or "",
        prev_model or "",
        m5_ctx or "",
        h4_ctx or "",
        hh_v or "",
        lh_v or "",
        ll_v or "",
        hl_v or "",
    ]
    ws.append(row)
    r = ws.max_row
    ws.cell(r, 1).number_format = "yyyy-mm-dd hh:mm:ss"

    saved = False
    try:
        wb.save(excel_path)
        _save_state(state_path, state)
        saved = True
        log.info(
            "models_excel ok tf=%s excel=%s model=%s HH=%s LH=%s LL=%s HL=%s",
            tf_label,
            excel_path,
            current_model or "(pusty)",
            hh_v or "-",
            lh_v or "-",
            ll_v or "-",
            hl_v or "-",
        )
    except PermissionError:
        log.warning(
            "models_excel: zamknij Excela lub plik zablokowany — nie zapisano %s", excel_path
        )
    except Exception:
        log.exception("models_excel: zapis Excel nieudany %s", excel_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return saved


def _setup_cli_logging() -> None:
    """Przy uruchomieniu z CLI zapis do pliku (pythonw/harmonogram ma osobny log w scheduled)."""
    log_path = Path(__file__).resolve().parent / "models_excel_cli.log"
    _LOG.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    has_file = any(
        isinstance(h, logging.FileHandler)
        and getattr(h, "baseFilename", None) == str(log_path.resolve())
        for h in _LOG.handlers
    )
    if not has_file:
        fh = logging.FileHandler(log_path, encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(fmt)
        _LOG.addHandler(fh)
    if sys.stderr and getattr(sys.stderr, "isatty", lambda: False)():
        has_sh = any(isinstance(h, logging.StreamHandler) for h in _LOG.handlers)
        if not has_sh:
            sh = logging.StreamHandler(sys.stderr)
            sh.setLevel(logging.INFO)
            sh.setFormatter(fmt)
            _LOG.addHandler(sh)
    _LOG.info("START models_excel_logger CLI log=%s", log_path.resolve())


def main() -> int:
    import argparse

    _setup_cli_logging()
    p = argparse.ArgumentParser(description="Test OCR + Excel (jeden PNG).")
    p.add_argument("--png", required=True, type=Path)
    p.add_argument("--tf", required=True, help="np. M5")
    p.add_argument("--repo-root", type=Path, default=None)
    p.add_argument("--config-json", type=Path, default=None, help="Fragment models_log z capture_schedule.json")
    args = p.parse_args()
    repo = args.repo_root or Path(__file__).resolve().parents[2]
    cfg: dict[str, Any] = {
        "excel_path": "models_logs/trading_models.xlsx",
        "schema_template": "models_logs/schema_template_tv_models.xlsx",
        "tesseract_cmd": "",
    }
    if args.config_json and args.config_json.is_file():
        cfg.update(json.loads(args.config_json.read_text(encoding="utf-8")))
    ts = datetime.now(_TS_PL).replace(microsecond=0).isoformat()
    try:
        ok = log_capture_to_excel(
            repo_root=repo,
            models_cfg=cfg,
            tf_label=args.tf,
            png_path=args.png,
            timestamp_pl_iso=ts,
            parent_log=_LOG,
        )
    except Exception:
        _LOG.exception("models_excel_logger CLI FATAL")
        return 1
    if not ok:
        _LOG.error(
            "models_excel_logger CLI — brak zapisu Excel (np. brak Tesseract lub zly PNG). "
            "Szczegoly w tym samym pliku logu."
        )
        print(
            "BLAD: nie zapisano trading_models.xlsx — zobacz scripts/tv_screenshot/models_excel_cli.log",
            file=sys.stderr,
        )
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
